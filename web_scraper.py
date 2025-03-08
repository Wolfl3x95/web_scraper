import requests
from bs4 import BeautifulSoup
import pandas as pd
from fpdf import FPDF
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

print("\n===== Web Scraper by W3X - Raffaele Brancaccio =====\n")

URL = input("Inserisci l'URL del sito da cui estrarre i dati: ")
format_choice = input("In quale formato vuoi salvare i dati? (csv/pdf/excel): ").strip().lower()

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}

currency_symbols = [
    "€", "$", "£", "¥", "₹", "₩", "₱", "₽", "฿", "₫", "₴", "₦", "₪", "₡", "₲", "₵", "₺"
]

response = requests.get(URL, headers=HEADERS)
if response.status_code != 200:
    print(f"Errore nella richiesta HTTP: {response.status_code}")
    exit()

soup = BeautifulSoup(response.text, "html.parser")

print("Anteprima della struttura HTML della pagina:")
print(soup.prettify()[:2000]) 

products = []
for item in soup.find_all(['div', 'article', 'li']): 
    title_element = item.find(['h1', 'h2', 'h3', 'h4', 'h5', 'h6']) 
    price_element = item.find(
        ['span', 'div', 'p'], 
        string=lambda text: text and any(c.isdigit() for c in text) and any(symbol in text for symbol in currency_symbols)
    ) 
    
    if title_element and price_element:
        title = title_element.get_text(strip=True)
        price = price_element.get_text(strip=True)
        products.append({"Titolo": title, "Prezzo": price})

if products:
    df = pd.DataFrame(products)
    
    if format_choice == "csv":
        df.to_csv("prodotti_estratti.csv", index=False)
        print("✅ Dati salvati con successo in prodotti_estratti.csv")
    elif format_choice == "pdf":
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, "Elenco Prodotti Estratti da Web Scraper by W3X - Raffaele Brancaccio", ln=True, align='C')
        pdf.ln(10)
        
        for product in products:
            title = product['Titolo'].encode('latin-1', 'ignore').decode('latin-1')
            price = product['Prezzo'].encode('latin-1', 'ignore').decode('latin-1')
            pdf.set_font("Arial", size=10)
            pdf.multi_cell(190, 8, f"Titolo: {title}")
            pdf.set_font("Arial", style='B', size=10)
            pdf.multi_cell(190, 8, f"Prezzo: {price}")
            pdf.ln(5)
            pdf.cell(190, 0, '', ln=True, border='B')
            pdf.ln(5)
        
        pdf.ln(10)
        pdf.set_font("Arial", style='B', size=12)
        pdf.cell(200, 10, "Creato da W3X - Raffaele Brancaccio", ln=True, align='C')
        pdf.cell(200, 10, "LinkedIn: Raffaele Brancaccio", ln=True, align='C')
        pdf.cell(200, 10, "GitHub: Wolfl3x95", ln=True, align='C')
        pdf.cell(200, 10, "Email: dev.raffaelebrancaccio@gmail.com", ln=True, align='C')
        
        pdf.output("prodotti_estratti.pdf")
        print("✅ Dati salvati con successo in prodotti_estratti.pdf")
    elif format_choice == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Elenco Prodotti Estratti da Web Scraper by W3X - Raffaele Brancaccio"
        
        headers = ["Titolo", "Prezzo"]
        ws.append(headers)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        for col_num, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for product in products:
            title_cell = ws.append([product["Titolo"], product["Prezzo"]])
            ws.cell(row=ws.max_row, column=2).font = Font(bold=True) 
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        
        ws.append([])
        bold_large_font = Font(bold=True, size=14)
        ws.append(["Creato da W3X - Raffaele Brancaccio"])
        ws.append(["LinkedIn: Raffaele Brancaccio"])
        ws.append(["GitHub: Wolfl3x95"])
        ws.append(["Email: dev.raffaelebrancaccio@gmail.com"])
        
        for row in range(ws.max_row - 3, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            cell.font = bold_large_font
            cell.alignment = Alignment(horizontal="center")
        
        wb.save("prodotti_estratti.xlsx")
        print("✅ Dati salvati con successo in prodotti_estratti.xlsx con tabella formattata.")
    else:
        print("⚠️ Formato non supportato. Usa csv, pdf o excel.")
else:
    print("⚠️ Nessun prodotto trovato. Controlla la struttura HTML stampata per verificare i tag corretti.")
